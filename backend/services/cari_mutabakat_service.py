"""
LYNTOS - Cari Mutabakat Servisi (IS-5)

SMMM'nin müşteri/tedarikçi cari hesap ekstrelerini yükleyip
mizan 120 (Alıcılar) ve 320 (Satıcılar) alt hesapları ile karşılaştırır.

Mevzuat Referansları:
- VUK Md. 177: Bilanço esasına göre defter tutma yükümlülüğü
- TTK Md. 64: Envanter çıkarma zorunluluğu
- VUK Md. 323: Şüpheli ticari alacaklar karşılığı (>365 gün)

TDHP Hesapları:
- 120 Alıcılar (120.001 Müşteri A, 120.002 Müşteri B...)
- 320 Satıcılar (320.001 Tedarikçi X, 320.002 Tedarikçi Y...)
- 128 Şüpheli Ticari Alacaklar
- 129 Şüpheli Alacak Karşılığı (-)

Mock/Demo Veri: YASAK
Sessiz Hata: YASAK - tüm hatalar loglanır
"""

import logging
import csv
import io
import re
from typing import List, Dict, Optional, Tuple
from datetime import datetime

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from database.db import get_connection

logger = logging.getLogger(__name__)

# Tolerans: 10 TL altı farklar "uyumlu" sayılır (sektör pratiği)
TOLERANS_TL = 10.0

# ═══════════════════════════════════════════════════════════
# SMMM KARAR SİSTEMİ (Pencere 7)
# ═══════════════════════════════════════════════════════════

def _ensure_smmm_kararlar_table():
    """smmm_kararlar tablosunu oluştur (yoksa)."""
    try:
        with get_connection() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS smmm_kararlar (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    client_id TEXT NOT NULL,
                    period_id TEXT NOT NULL,
                    hesap_kodu TEXT NOT NULL,
                    karar TEXT NOT NULL,
                    smmm_id TEXT NOT NULL DEFAULT 'SMMM',
                    not_metni TEXT DEFAULT '',
                    created_at TEXT DEFAULT (datetime('now')),
                    updated_at TEXT DEFAULT (datetime('now')),
                    UNIQUE(client_id, period_id, hesap_kodu)
                )
            """)
            conn.commit()
    except Exception as e:
        logger.error(f"[SmmmKarar] Tablo oluşturma hatası: {e}")


def save_karar(
    client_id: str,
    period_id: str,
    hesap_kodu: str,
    karar: str,
    smmm_id: str = "SMMM",
    not_metni: str = "",
) -> Dict:
    """Tek karar kaydet (UPSERT)."""
    _ensure_smmm_kararlar_table()

    valid_kararlar = ("RESMI", "DEFTER_DISI", "BILINMIYOR")
    if karar not in valid_kararlar:
        raise ValueError(f"Geçersiz karar: {karar}. Geçerli: {valid_kararlar}")

    with get_connection() as conn:
        conn.execute("""
            INSERT INTO smmm_kararlar (client_id, period_id, hesap_kodu, karar, smmm_id, not_metni, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(client_id, period_id, hesap_kodu)
            DO UPDATE SET karar=excluded.karar, smmm_id=excluded.smmm_id,
                          not_metni=excluded.not_metni, updated_at=datetime('now')
        """, (client_id, period_id, hesap_kodu, karar, smmm_id, not_metni))
        conn.commit()

    logger.info(f"[SmmmKarar] Karar kaydedildi: {hesap_kodu} → {karar} ({client_id}/{period_id})")
    return {
        "hesap_kodu": hesap_kodu,
        "karar": karar,
        "smmm_id": smmm_id,
        "not_metni": not_metni,
    }


def save_kararlar_toplu(
    client_id: str,
    period_id: str,
    kararlar: List[Dict],
    smmm_id: str = "SMMM",
) -> Dict:
    """Toplu karar kaydet."""
    _ensure_smmm_kararlar_table()

    valid_kararlar = ("RESMI", "DEFTER_DISI", "BILINMIYOR")
    saved = 0

    with get_connection() as conn:
        for item in kararlar:
            hesap_kodu = item.get("hesap_kodu", "")
            karar = item.get("karar", "")
            not_metni = item.get("not_metni", item.get("not", ""))

            if not hesap_kodu or karar not in valid_kararlar:
                continue

            conn.execute("""
                INSERT INTO smmm_kararlar (client_id, period_id, hesap_kodu, karar, smmm_id, not_metni, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(client_id, period_id, hesap_kodu)
                DO UPDATE SET karar=excluded.karar, smmm_id=excluded.smmm_id,
                              not_metni=excluded.not_metni, updated_at=datetime('now')
            """, (client_id, period_id, hesap_kodu, karar, smmm_id, not_metni))
            saved += 1

        conn.commit()

    logger.info(f"[SmmmKarar] Toplu karar: {saved}/{len(kararlar)} kaydedildi ({client_id}/{period_id})")
    return {"saved": saved, "total": len(kararlar)}


def get_kararlar(client_id: str, period_id: str) -> Dict[str, Dict]:
    """Tüm kararları getir. Returns: {hesap_kodu: {karar, not, tarih}}"""
    _ensure_smmm_kararlar_table()

    result = {}
    with get_connection() as conn:
        rows = conn.execute("""
            SELECT hesap_kodu, karar, not_metni, updated_at
            FROM smmm_kararlar
            WHERE client_id = ? AND period_id = ?
            ORDER BY hesap_kodu
        """, (client_id, period_id)).fetchall()

        for row in rows:
            result[row["hesap_kodu"]] = {
                "karar": row["karar"],
                "not": row["not_metni"] or "",
                "tarih": row["updated_at"] or "",
            }

    return result

# VUK Md. 323: 365 günü aşan alacaklar için şüpheli alacak karşılığı
SUPHELI_ALACAK_GUN_ESIGI = 365


def get_mizan_cari_hesaplar(client_id: str, period_id: str) -> Dict[str, Dict]:
    """
    Mizan'dan 120.xxx (Alıcılar) ve 320.xxx (Satıcılar) alt hesaplarını çeker.

    120 grubu: Borç bakiye = müşteriden alacak
    320 grubu: Alacak bakiye = tedarikçiye borç

    Returns:
        Dict[hesap_kodu, {hesap_adi, borc_bakiye, alacak_bakiye, net_bakiye}]
    """
    result = {}
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT hesap_kodu, hesap_adi, borc_bakiye, alacak_bakiye
                FROM mizan_entries
                WHERE client_id = ? AND period_id = ?
                  AND (hesap_kodu LIKE '120.%' OR hesap_kodu LIKE '320.%'
                       OR hesap_kodu LIKE '120,%' OR hesap_kodu LIKE '320,%')
                ORDER BY hesap_kodu
            """, (client_id, period_id))

            rows = cursor.fetchall()
            for row in rows:
                hesap_kodu = row['hesap_kodu']
                borc = row['borc_bakiye'] or 0.0
                alacak = row['alacak_bakiye'] or 0.0

                # 120 grubu: borç bakiye (müşteriden alacak)
                # 320 grubu: alacak bakiye (tedarikçiye borç)
                if hesap_kodu.startswith('120'):
                    net_bakiye = borc - alacak  # Borç bakiye vermeli
                else:
                    net_bakiye = alacak - borc  # Alacak bakiye vermeli

                result[hesap_kodu] = {
                    'hesap_adi': row['hesap_adi'] or hesap_kodu,
                    'borc_bakiye': borc,
                    'alacak_bakiye': alacak,
                    'net_bakiye': net_bakiye,
                }

            logger.info(
                f"[CariMutabakat] Mizan'dan {len(result)} cari hesap çekildi "
                f"(client={client_id}, period={period_id})"
            )
    except Exception as e:
        logger.error(f"[CariMutabakat] Mizan cari hesap çekme hatası: {e}")
        raise

    return result


def parse_ekstre_csv(content: bytes, filename: str) -> List[Dict]:
    """
    Cari hesap ekstre CSV/TXT dosyasını parse eder.

    Beklenen CSV formatı (başlık satırı zorunlu):
    hesap_kodu;karsi_taraf;bakiye

    Alternatif ayırıcılar: ; , \\t
    Türkçe ondalık: 1.234,56 → 1234.56

    Returns:
        List[{hesap_kodu, karsi_taraf, bakiye}]
    """
    rows = []
    try:
        # Encoding denemeleri: UTF-8, Windows-1254 (Türkçe), Latin-1
        text = None
        for encoding in ['utf-8-sig', 'utf-8', 'cp1254', 'latin-1']:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if text is None:
            logger.error(f"[CariMutabakat] Dosya encode edilemedi: {filename}")
            raise ValueError("Dosya karakter kodlaması tanınamadı")

        # Ayırıcı tespiti
        first_line = text.strip().split('\n')[0] if text.strip() else ''
        if ';' in first_line:
            delimiter = ';'
        elif '\t' in first_line:
            delimiter = '\t'
        else:
            delimiter = ','

        reader = csv.DictReader(
            io.StringIO(text),
            delimiter=delimiter,
        )

        # Başlık normalizasyonu (küçük harf, boşluk temizle)
        if reader.fieldnames:
            reader.fieldnames = [f.strip().lower().replace(' ', '_') for f in reader.fieldnames]

        for i, row in enumerate(reader):
            # Hesap kodu bul
            hesap_kodu = (
                row.get('hesap_kodu') or row.get('hesap') or
                row.get('account_code') or row.get('kod') or ''
            ).strip()

            if not hesap_kodu:
                logger.warning(f"[CariMutabakat] Satır {i+2}: hesap_kodu boş, atlanıyor")
                continue

            # Karşı taraf unvanı
            karsi_taraf = (
                row.get('karsi_taraf') or row.get('karsi_taraf_unvan') or
                row.get('unvan') or row.get('firma') or
                row.get('musteri') or row.get('tedarikci') or ''
            ).strip()

            # Bakiye parse (Türkçe format: 1.234,56 → 1234.56)
            bakiye_str = (
                row.get('bakiye') or row.get('ekstre_bakiye') or
                row.get('balance') or row.get('tutar') or '0'
            ).strip()
            bakiye = _parse_turkish_number(bakiye_str)

            rows.append({
                'hesap_kodu': hesap_kodu,
                'karsi_taraf': karsi_taraf,
                'bakiye': bakiye,
            })

        logger.info(f"[CariMutabakat] CSV parse: {len(rows)} satır okundu ({filename})")

    except ValueError:
        raise
    except Exception as e:
        logger.error(f"[CariMutabakat] CSV parse hatası ({filename}): {e}")
        raise ValueError(f"CSV dosyası okunamadı: {e}")

    return rows


def parse_ekstre_excel(content: bytes, filename: str) -> List[Dict]:
    """
    Cari hesap ekstre Excel dosyasını parse eder (.xlsx / .xls).

    Beklenen sütunlar: hesap_kodu, karsi_taraf/unvan, bakiye

    Returns:
        List[{hesap_kodu, karsi_taraf, bakiye}]
    """
    rows = []
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel dosyasında aktif sayfa bulunamadı")

        # Başlıkları oku (ilk satır)
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            val = str(cell.value or '').strip().lower().replace(' ', '_')
            headers.append(val)

        # Hesap kodu, karşı taraf, bakiye sütun indekslerini bul
        hesap_idx = _find_column_index(headers, ['hesap_kodu', 'hesap', 'kod', 'account_code'])
        karsi_idx = _find_column_index(headers, ['karsi_taraf', 'karsi_taraf_unvan', 'unvan', 'firma', 'musteri', 'tedarikci'])
        bakiye_idx = _find_column_index(headers, ['bakiye', 'ekstre_bakiye', 'balance', 'tutar'])

        if hesap_idx is None:
            raise ValueError("Excel'de 'hesap_kodu' sütunu bulunamadı")
        if bakiye_idx is None:
            raise ValueError("Excel'de 'bakiye' sütunu bulunamadı")

        for row_cells in ws.iter_rows(min_row=2):
            cells = list(row_cells)
            hesap_kodu = str(cells[hesap_idx].value or '').strip()
            if not hesap_kodu:
                continue

            karsi_taraf = ''
            if karsi_idx is not None and karsi_idx < len(cells):
                karsi_taraf = str(cells[karsi_idx].value or '').strip()

            bakiye_val = cells[bakiye_idx].value
            if isinstance(bakiye_val, (int, float)):
                bakiye = float(bakiye_val)
            else:
                bakiye = _parse_turkish_number(str(bakiye_val or '0'))

            rows.append({
                'hesap_kodu': hesap_kodu,
                'karsi_taraf': karsi_taraf,
                'bakiye': bakiye,
            })

        wb.close()
        logger.info(f"[CariMutabakat] Excel parse: {len(rows)} satır okundu ({filename})")

    except ImportError:
        logger.error("[CariMutabakat] openpyxl kütüphanesi yüklü değil")
        raise ValueError("Excel desteği için openpyxl gerekli. pip install openpyxl")
    except ValueError:
        raise
    except Exception as e:
        logger.error(f"[CariMutabakat] Excel parse hatası ({filename}): {e}")
        raise ValueError(f"Excel dosyası okunamadı: {e}")

    return rows


def run_mutabakat(
    client_id: str,
    period_id: str,
    ekstre_rows: List[Dict],
    source_file: str,
) -> Dict:
    """
    Cari mutabakat hesaplama:
    1. Mizan'dan 120.xxx / 320.xxx alt hesapları çek
    2. Ekstre bakiyeleri ile karşılaştır
    3. Tolerans: 10 TL altı fark = "uyumlu"
    4. Aging: >365 gün → şüpheli alacak uyarısı (VUK 323)
    5. DB'ye kaydet

    Returns:
        {sonuclar: List, ozet: Dict}
    """
    mizan_hesaplar = get_mizan_cari_hesaplar(client_id, period_id)
    sonuclar = []

    try:
        with get_connection() as conn:
            cursor = conn.cursor()

            # Önceki mutabakat verilerini temizle (aynı dönem, aynı dosya)
            cursor.execute("""
                DELETE FROM cari_ekstreler
                WHERE client_id = ? AND period_id = ?
            """, (client_id, period_id))

            for ekstre in ekstre_rows:
                hesap_kodu = ekstre['hesap_kodu']
                karsi_taraf = ekstre.get('karsi_taraf', '')
                ekstre_bakiye = ekstre['bakiye']

                # Mizan karşılaştırması
                mizan_info = mizan_hesaplar.get(hesap_kodu)
                if mizan_info:
                    mizan_bakiye = mizan_info['net_bakiye']
                else:
                    mizan_bakiye = 0.0
                    logger.warning(
                        f"[CariMutabakat] Mizan'da hesap bulunamadı: {hesap_kodu} "
                        f"(client={client_id}, period={period_id})"
                    )

                fark = abs(ekstre_bakiye - mizan_bakiye)
                fark_yuzde = 0.0
                if mizan_bakiye != 0:
                    fark_yuzde = round((fark / abs(mizan_bakiye)) * 100, 2)

                # Durum belirleme: 10 TL tolerans
                if fark <= TOLERANS_TL:
                    durum = 'uyumlu'
                else:
                    durum = 'farkli'

                # Aging ve şüpheli alacak kontrolü (VUK Md. 323)
                # Sadece 120 grubu (alıcılar) için aging kontrolü
                aging_gun = 0
                supheli_riski = 0
                if hesap_kodu.startswith('120') and mizan_bakiye > 0:
                    aging_gun = _hesapla_aging(client_id, period_id, hesap_kodu)
                    if aging_gun > SUPHELI_ALACAK_GUN_ESIGI:
                        supheli_riski = 1

                # DB'ye kaydet
                cursor.execute("""
                    INSERT INTO cari_ekstreler
                    (client_id, period_id, cari_hesap_kodu, karsi_taraf,
                     ekstre_bakiye, mizan_bakiye, fark, fark_yuzde,
                     durum, aging_gun, supheli_alacak_riski, source_file)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    client_id, period_id, hesap_kodu, karsi_taraf,
                    ekstre_bakiye, mizan_bakiye, fark, fark_yuzde,
                    durum, aging_gun, supheli_riski, source_file,
                ))

                sonuclar.append({
                    'hesap_kodu': hesap_kodu,
                    'hesap_adi': mizan_info['hesap_adi'] if mizan_info else karsi_taraf or hesap_kodu,
                    'karsi_taraf': karsi_taraf,
                    'ekstre_bakiye': ekstre_bakiye,
                    'mizan_bakiye': mizan_bakiye,
                    'fark': fark,
                    'fark_yuzde': fark_yuzde,
                    'durum': durum,
                    'aging_gun': aging_gun,
                    'supheli_alacak_riski': bool(supheli_riski),
                })

            conn.commit()

        # Mizan'da olup ekstre'de olmayan hesapları da raporla
        ekstre_kodlar = {e['hesap_kodu'] for e in ekstre_rows}
        for hesap_kodu, info in mizan_hesaplar.items():
            if hesap_kodu not in ekstre_kodlar and abs(info['net_bakiye']) > TOLERANS_TL:
                sonuclar.append({
                    'hesap_kodu': hesap_kodu,
                    'hesap_adi': info['hesap_adi'],
                    'karsi_taraf': '',
                    'ekstre_bakiye': 0.0,
                    'mizan_bakiye': info['net_bakiye'],
                    'fark': abs(info['net_bakiye']),
                    'fark_yuzde': 100.0,
                    'durum': 'farkli',
                    'aging_gun': 0,
                    'supheli_alacak_riski': False,
                    'not': 'Ekstre yüklenmemiş - sadece mizan bakiyesi mevcut',
                })

        # Özet hesapla
        toplam = len(sonuclar)
        uyumlu = sum(1 for s in sonuclar if s['durum'] == 'uyumlu')
        farkli = sum(1 for s in sonuclar if s['durum'] == 'farkli')
        toplam_fark = sum(s['fark'] for s in sonuclar)
        supheli_sayisi = sum(1 for s in sonuclar if s.get('supheli_alacak_riski'))

        ozet = {
            'toplam_hesap': toplam,
            'uyumlu_hesap': uyumlu,
            'farkli_hesap': farkli,
            'toplam_fark': round(toplam_fark, 2),
            'supheli_alacak_sayisi': supheli_sayisi,
            'tolerans_tl': TOLERANS_TL,
            'mevzuat_ref': 'VUK Md. 177, TTK Md. 64, VUK Md. 323',
        }

        logger.info(
            f"[CariMutabakat] Mutabakat tamamlandı: "
            f"{toplam} hesap, {uyumlu} uyumlu, {farkli} farklı, "
            f"toplam fark: {toplam_fark:,.2f} TL "
            f"(client={client_id}, period={period_id})"
        )

        return {'sonuclar': sonuclar, 'ozet': ozet}

    except Exception as e:
        logger.error(f"[CariMutabakat] Mutabakat hesaplama hatası: {e}")
        raise


def get_mutabakat_list(
    client_id: str,
    period_id: str,
    filtre: Optional[str] = None,
) -> List[Dict]:
    """
    Kayıtlı mutabakat sonuçlarını listeler.

    filtre: 'tumu' | 'farkli' | 'onaylanan' | 'supheli' | None
    """
    try:
        with get_connection() as conn:
            cursor = conn.cursor()

            query = """
                SELECT id, cari_hesap_kodu, karsi_taraf,
                       ekstre_bakiye, mizan_bakiye, fark, fark_yuzde,
                       durum, aging_gun, supheli_alacak_riski,
                       onaylayan, onay_tarihi, source_file, uploaded_at
                FROM cari_ekstreler
                WHERE client_id = ? AND period_id = ?
            """
            params: list = [client_id, period_id]

            if filtre == 'farkli':
                query += " AND durum = 'farkli'"
            elif filtre == 'onaylanan':
                query += " AND durum = 'onaylandi'"
            elif filtre == 'supheli':
                query += " AND supheli_alacak_riski = 1"

            query += " ORDER BY fark DESC"

            cursor.execute(query, params)
            rows = cursor.fetchall()

            result = []
            for row in rows:
                result.append({
                    'id': row['id'],
                    'hesap_kodu': row['cari_hesap_kodu'],
                    'karsi_taraf': row['karsi_taraf'] or '',
                    'ekstre_bakiye': row['ekstre_bakiye'],
                    'mizan_bakiye': row['mizan_bakiye'],
                    'fark': row['fark'],
                    'fark_yuzde': row['fark_yuzde'],
                    'durum': row['durum'],
                    'aging_gun': row['aging_gun'],
                    'supheli_alacak_riski': bool(row['supheli_alacak_riski']),
                    'onaylayan': row['onaylayan'],
                    'onay_tarihi': row['onay_tarihi'],
                    'source_file': row['source_file'],
                    'uploaded_at': row['uploaded_at'],
                })

            logger.info(
                f"[CariMutabakat] Liste: {len(result)} kayıt "
                f"(client={client_id}, period={period_id}, filtre={filtre})"
            )
            return result

    except Exception as e:
        logger.error(f"[CariMutabakat] Liste çekme hatası: {e}")
        raise


def onayla_mutabakat(
    ids: List[int],
    onaylayan: str,
) -> Dict:
    """
    Seçili mutabakat satırlarını SMMM onaylar.
    Durum: 'farkli' veya 'uyumlu' → 'onaylandi'
    """
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            onay_tarihi = datetime.now().isoformat()

            updated = 0
            for row_id in ids:
                cursor.execute("""
                    UPDATE cari_ekstreler
                    SET durum = 'onaylandi',
                        onaylayan = ?,
                        onay_tarihi = ?,
                        updated_at = datetime('now')
                    WHERE id = ?
                """, (onaylayan, onay_tarihi, row_id))
                updated += cursor.rowcount

            conn.commit()

            logger.info(
                f"[CariMutabakat] {updated}/{len(ids)} satır onaylandı "
                f"(onaylayan={onaylayan})"
            )
            return {
                'onaylanan_sayisi': updated,
                'onaylayan': onaylayan,
                'onay_tarihi': onay_tarihi,
            }

    except Exception as e:
        logger.error(f"[CariMutabakat] Onay hatası: {e}")
        raise


def get_mutabakat_ozet(client_id: str, period_id: str) -> Dict:
    """
    Dashboard için mutabakat özeti.
    """
    try:
        with get_connection() as conn:
            cursor = conn.cursor()

            # Toplam sayılar
            cursor.execute("""
                SELECT
                    COUNT(*) as toplam,
                    SUM(CASE WHEN durum = 'uyumlu' THEN 1 ELSE 0 END) as uyumlu,
                    SUM(CASE WHEN durum = 'farkli' THEN 1 ELSE 0 END) as farkli,
                    SUM(CASE WHEN durum = 'onaylandi' THEN 1 ELSE 0 END) as onaylanan,
                    SUM(CASE WHEN durum = 'beklemede' THEN 1 ELSE 0 END) as beklemede,
                    SUM(fark) as toplam_fark,
                    SUM(CASE WHEN supheli_alacak_riski = 1 THEN 1 ELSE 0 END) as supheli_sayisi,
                    MAX(uploaded_at) as son_yukleme
                FROM cari_ekstreler
                WHERE client_id = ? AND period_id = ?
            """, (client_id, period_id))

            row = cursor.fetchone()

            if row['toplam'] == 0:
                return {
                    'veri_var': False,
                    'toplam_hesap': 0,
                    'uyumlu': 0,
                    'farkli': 0,
                    'onaylanan': 0,
                    'beklemede': 0,
                    'toplam_fark': 0.0,
                    'supheli_alacak_sayisi': 0,
                    'son_yukleme': None,
                    'mevzuat_ref': 'VUK Md. 177, TTK Md. 64',
                }

            return {
                'veri_var': True,
                'toplam_hesap': row['toplam'],
                'uyumlu': row['uyumlu'] or 0,
                'farkli': row['farkli'] or 0,
                'onaylanan': row['onaylanan'] or 0,
                'beklemede': row['beklemede'] or 0,
                'toplam_fark': round(row['toplam_fark'] or 0, 2),
                'supheli_alacak_sayisi': row['supheli_sayisi'] or 0,
                'son_yukleme': row['son_yukleme'],
                'mevzuat_ref': 'VUK Md. 177, TTK Md. 64, VUK Md. 323',
            }

    except Exception as e:
        logger.error(f"[CariMutabakat] Özet çekme hatası: {e}")
        raise


# ═══════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════

def _parse_turkish_number(s: str) -> float:
    """
    Türkçe sayı formatını Python float'a çevirir.
    1.234.567,89 → 1234567.89
    1234,56 → 1234.56
    -1.234,56 → -1234.56
    """
    if not s or s.strip() == '':
        return 0.0
    s = s.strip()

    # Negatif işareti
    negative = s.startswith('-') or s.startswith('(')
    s = s.replace('-', '').replace('(', '').replace(')', '')

    # Binlik ayırıcı nokta, ondalık virgül
    if ',' in s and '.' in s:
        # 1.234,56 formatı
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        # 1234,56 formatı
        s = s.replace(',', '.')
    # else: zaten 1234.56 veya düz sayı

    try:
        val = float(s)
        return -val if negative else val
    except ValueError:
        logger.warning(f"[CariMutabakat] Sayı parse edilemedi: '{s}'")
        return 0.0


def _find_column_index(headers: List[str], candidates: List[str]) -> Optional[int]:
    """Excel başlık listesinde aday sütun isimlerinden birini bul."""
    for candidate in candidates:
        for i, h in enumerate(headers):
            if h == candidate:
                return i
    return None


def _hesapla_aging(client_id: str, period_id: str, hesap_kodu: str) -> int:
    """
    Hesap için aging (yaşlandırma) gün sayısı hesapla.

    Yevmiye kayıtlarından son hareket tarihini bulur.
    Hareket yoksa dönem başından itibaren sayar.

    VUK Md. 323: >365 gün → şüpheli alacak karşılığı gerekir.
    """
    try:
        with get_connection() as conn:
            cursor = conn.cursor()

            # Yevmiye'den son hareket tarihi
            cursor.execute("""
                SELECT MAX(tarih) as son_hareket
                FROM journal_entries
                WHERE client_id = ? AND period_id = ? AND hesap_kodu = ?
            """, (client_id, period_id, hesap_kodu))

            row = cursor.fetchone()
            son_hareket = row['son_hareket'] if row else None

            if son_hareket:
                try:
                    # Tarih formatları: YYYY-MM-DD veya DD.MM.YYYY
                    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                        try:
                            tarih = datetime.strptime(son_hareket, fmt)
                            gun_farki = (datetime.now() - tarih).days
                            return max(gun_farki, 0)
                        except ValueError:
                            continue
                except Exception:
                    pass

            # Hareket bulunamadıysa, dönem başından say
            # 2025-Q1 → 2025-01-01
            try:
                yil = int(period_id.split('-')[0])
                return (datetime.now() - datetime(yil, 1, 1)).days
            except Exception:
                return 0

    except Exception as e:
        logger.warning(f"[CariMutabakat] Aging hesaplama hatası ({hesap_kodu}): {e}")
        return 0


# ═══════════════════════════════════════════════════════════
# AKILLI CSV PARSER: SÖZLÜK + FUZZY MATCHING + PROGRAM TESPİTİ
# ═══════════════════════════════════════════════════════════

# Türkçe karakter normalizasyon tablosu
_TR_CHAR_MAP = str.maketrans('çğıöşüÇĞİÖŞÜ', 'cgiosuCGIOSU')


COLUMN_ALIASES: Dict[str, List[str]] = {
    'hesap_kodu': [
        # Genel
        'hesap_kodu', 'hesap_no', 'hesapno', 'hsp_kodu', 'hspkodu',
        'hkod', 'h_kodu', 'cari_hesap_kodu', 'cari_kod', 'cari_no',
        'musteri_kodu', 'tedarikci_kodu', 'kod', 'no',
        'hesap_numarasi', 'hesapnumarasi',
        # Logo Tiger / Unity
        'hesap kodu', 'hesapkodu', 'cari hesap kodu', 'cari hesap no',
        'hesap numarası', 'hesap no.', 'hs. kodu',
        # Mikro
        'hsp kodu', 'hesno', 'hsp no', 'hsp. kodu', 'hsp.kodu',
        'muhasebe hesap kodu', 'muh. hesap kodu',
        # ETA / Eta SQL
        'hesap numarasi', 'h.kodu', 'hesap_no', 'cari hesap numarasi',
        # Luca
        'h_no', 'hesap_numarasi', 'cari_hesap_no',
        # Zirve
        'muhasebe_kodu', 'muh_kodu', 'muhasebe kodu', 'muh.kodu',
        # Netsis / Canias / SAP
        'gl_code', 'account_no', 'konto', 'konto_no', 'sachkonto',
        # İngilizce
        'account_code', 'acc_code', 'acct_code', 'account_number',
        'acc_no', 'gl_account',
    ],
    'karsi_taraf': [
        # Genel
        'karsi_taraf', 'karsi_taraf_unvan', 'unvan', 'firma',
        'firma_adi', 'firma_unvan', 'musteri', 'musteri_adi',
        'tedarikci', 'tedarikci_adi', 'isim', 'ad_soyad',
        'cari_unvan', 'cari_hesap_adi', 'hesap_adi', 'hesapadi',
        'sirket_adi', 'sirket', 'ticaret_unvani',
        # Logo Tiger / Unity
        'hesap adı', 'hesap adi', 'cari hesap adı', 'firma adı',
        'ünvan', 'ticaret ünvanı',
        # Mikro
        'firmaadi', 'firma adi', 'cari adi', 'cari adı',
        'unvani', 'ünvanı',
        # ETA
        'hesap ismi', 'hesap_ismi', 'cari isim', 'cari_isim',
        # Luca
        'hesap_adi', 'h_adi',
        # Zirve
        'firma_ad', 'cari_ad',
        # Netsis / SAP
        'name1', 'partner_name', 'bp_name',
        # İngilizce
        'account_name', 'company', 'name', 'vendor', 'customer',
        'counterparty', 'party_name', 'trading_partner',
    ],
    'bakiye': [
        # Genel
        'bakiye', 'ekstre_bakiye', 'net_bakiye', 'borc_bakiye',
        'alacak_bakiye', 'kalan', 'mutabakat_bakiye',
        'donem_sonu_bakiye', 'son_bakiye', 'tutar', 'toplam',
        'net_tutar', 'bakiye_tl',
        # Logo Tiger
        'bakiye', 'net bakiye', 'borç bakiye', 'alacak bakiye',
        'kalan bakiye', 'dönem sonu bakiye',
        # Mikro
        'bakiye_tl', 'net_tutar', 'tutar_tl',
        # ETA
        'bakiye tutarı', 'bakiye_tutari', 'net tutar',
        # Luca
        'donem_bakiye', 'net_bakiye',
        # Zirve
        'net tutar', 'kalan tutar',
        # Netsis
        'saldo', 'betrag',
        # İngilizce
        'balance', 'amount', 'total', 'net_balance',
        'net_amount', 'closing_balance', 'ending_balance',
    ],
    'borc': [
        'borc', 'borç', 'borc_toplam', 'borc_tutari', 'borç_toplamı',
        'borc toplam', 'borç toplam', 'debit', 'soll',
        'borc_bakiye', 'donem_borc',
    ],
    'alacak': [
        'alacak', 'alacak_toplam', 'alacak_tutari', 'alacak_toplamı',
        'alacak toplam', 'credit', 'haben',
        'alacak_bakiye', 'donem_alacak',
    ],
}

# Muhasebe programı header pattern'leri
_PROGRAM_SIGNATURES: Dict[str, Dict] = {
    'Logo Tiger': {
        'patterns': [
            ['hesap kodu', 'hesap adı', 'borç', 'alacak', 'bakiye'],
            ['hesap kodu', 'hesap adı', 'borç', 'alacak'],
            ['cari hesap kodu', 'ünvan', 'bakiye'],
        ],
        'delimiter': ';',
    },
    'Mikro': {
        'patterns': [
            ['hsp kodu', 'firma', 'tutar'],
            ['hspkodu', 'firmaadi', 'bakiye'],
            ['hsp.kodu', 'firma adı', 'net tutar'],
            ['hsp kodu', 'cari adı', 'bakiye'],
        ],
        'delimiter': '\t',
    },
    'ETA': {
        'patterns': [
            ['hesap numarası', 'ünvan', 'bakiye'],
            ['hesap numarası', 'ünvan', 'bakiye tutarı'],
            ['hesap numarasi', 'hesap ismi', 'bakiye tutarı'],
            ['h.kodu', 'hesap ismi', 'net tutar'],
        ],
        'delimiter': '|',
    },
    'Zirve': {
        'patterns': [
            ['muhasebe kodu', 'firma', 'net tutar'],
            ['muhasebe_kodu', 'firma_adi', 'bakiye'],
            ['muh.kodu', 'cari_ad', 'kalan tutar'],
        ],
        'delimiter': ',',
    },
    'Luca': {
        'patterns': [
            ['hesap_kodu', 'hesap_adi', 'borc', 'alacak'],
            ['h_no', 'h_adi', 'net_bakiye'],
            ['cari_hesap_no', 'cari_unvan', 'donem_bakiye'],
        ],
        'delimiter': ';',
    },
    'Netsis': {
        'patterns': [
            ['gl_code', 'name1', 'saldo'],
            ['konto', 'partner_name', 'betrag'],
            ['account_no', 'account_name', 'balance'],
        ],
        'delimiter': ';',
    },
}

# Regex pattern'ler (son çare eşleşme)
_REGEX_PATTERNS: Dict[str, List[str]] = {
    'hesap_kodu': [r'hesap.*kod', r'hsp.*kod', r'cari.*kod', r'cari.*no',
                   r'acc.*code', r'gl.*code', r'muhasebe.*kod', r'konto'],
    'karsi_taraf': [r'karsi.*taraf', r'cari.*unvan', r'firma.*ad', r'hesap.*ad',
                    r'cari.*ad', r'musteri.*ad', r'tedarikci.*ad', r'unvan',
                    r'account.*name', r'company', r'partner.*name'],
    'bakiye': [r'bak.*ye', r'net.*tutar', r'net.*bak', r'son.*bak',
               r'kalan', r'balance', r'amount', r'saldo', r'toplam'],
    'borc': [r'bor[cç]', r'debit', r'soll'],
    'alacak': [r'alacak', r'credit', r'haben'],
}


def _normalize_header(s: str) -> str:
    """Sütun başlığını normalize et: küçük harf, Türkçe karakter, boşluk birleştir."""
    s = s.strip().lower()
    s = s.translate(_TR_CHAR_MAP)
    # BOM temizle
    s = s.replace('\ufeff', '')
    # Boşluk ve alt çizgi birleştir
    s = re.sub(r'[\s_\-\.]+', '_', s)
    # Baş/son alt çizgi temizle
    s = s.strip('_')
    return s


def _levenshtein(a: str, b: str) -> int:
    """Basit Levenshtein mesafesi hesaplama."""
    if len(a) < len(b):
        return _levenshtein(b, a)
    if len(b) == 0:
        return len(a)
    prev_row = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        curr_row = [i + 1]
        for j, cb in enumerate(b):
            cost = 0 if ca == cb else 1
            curr_row.append(min(
                curr_row[j] + 1,
                prev_row[j + 1] + 1,
                prev_row[j] + cost,
            ))
        prev_row = curr_row
    return prev_row[-1]


def _match_column(raw_header: str, target_field: str) -> Tuple[bool, int, str]:
    """
    Tek bir raw header'ı target_field'a eşleştirmeye çalışır.

    Returns: (matched, confidence, method)
        method: 'exact', 'fuzzy', 'pattern'
    """
    normalized = _normalize_header(raw_header)
    aliases = COLUMN_ALIASES.get(target_field, [])

    # 1. Tam eşleşme (100%)
    for alias in aliases:
        norm_alias = _normalize_header(alias)
        if normalized == norm_alias:
            return True, 100, 'exact'

    # 2. Fuzzy (Levenshtein ≤ 2) → 70-90%
    for alias in aliases:
        norm_alias = _normalize_header(alias)
        if len(norm_alias) < 3:
            continue  # Çok kısa alias'larda fuzzy yapma
        dist = _levenshtein(normalized, norm_alias)
        if dist <= 2 and dist < len(norm_alias) * 0.4:
            confidence = max(70, 100 - dist * 15)
            return True, confidence, 'fuzzy'

    # 3. Regex pattern (50-70%)
    patterns = _REGEX_PATTERNS.get(target_field, [])
    for pat in patterns:
        if re.search(pat, normalized):
            return True, 60, 'pattern'

    return False, 0, ''


def smart_map_columns(raw_headers: List[str]) -> Dict[str, Dict]:
    """
    Raw CSV/Excel başlıklarını canonical alanlara akıllı eşleştirir.

    Returns: {
        'hesap_kodu': {'source_column': 'HspKodu', 'source_index': 0, 'confidence': 92, 'method': 'fuzzy'},
        'karsi_taraf': {'source_column': 'Firma Adı', 'source_index': 1, 'confidence': 88, 'method': 'exact'},
        ...
    }
    """
    mapping: Dict[str, Dict] = {}
    target_fields = ['hesap_kodu', 'karsi_taraf', 'bakiye', 'borc', 'alacak']

    for field in target_fields:
        best_confidence = 0
        best_idx = -1
        best_method = ''
        best_raw = ''

        for idx, raw in enumerate(raw_headers):
            matched, confidence, method = _match_column(raw, field)
            if matched and confidence > best_confidence:
                best_confidence = confidence
                best_idx = idx
                best_method = method
                best_raw = raw

        if best_confidence > 0:
            mapping[field] = {
                'source_column': best_raw,
                'source_index': best_idx,
                'confidence': best_confidence,
                'method': best_method,
            }

    # bakiye yoksa ama borc+alacak varsa, computed bakiye olarak işaretle
    if 'bakiye' not in mapping and 'borc' in mapping and 'alacak' in mapping:
        mapping['bakiye'] = {
            'source_column': f"{mapping['borc']['source_column']} - {mapping['alacak']['source_column']}",
            'source_index': -1,  # computed
            'confidence': 85,
            'method': 'computed',
            'computed_from': ['borc', 'alacak'],
        }

    return mapping


def detect_accounting_program(raw_headers: List[str], delimiter: str) -> Optional[str]:
    """
    CSV header ve delimiter'a göre muhasebe programını tespit et.

    Returns: Program adı veya None
    """
    normalized_headers = [_normalize_header(h) for h in raw_headers]

    for program, sig in _PROGRAM_SIGNATURES.items():
        for pattern_set in sig['patterns']:
            normalized_pattern = [_normalize_header(p) for p in pattern_set]
            matched = sum(1 for np in normalized_pattern if np in normalized_headers)
            # Pattern'in %70'i eşleşirse program tespit edildi
            if matched >= len(normalized_pattern) * 0.7:
                # Delimiter de uyuşuyorsa confidence artır
                if delimiter == sig.get('delimiter', ','):
                    return program
                else:
                    return program  # Delimiter farklı olsa bile header eşleşmesi yeterli

    return None


def preview_ekstre(content: bytes, filename: str) -> Dict:
    """
    CSV/Excel dosyasını parse edip preview bilgisi döndürür.
    Mutabakat çalıştırmaz, sadece sütun haritası ve örnek satırlar döner.

    Returns: {
        'column_mapping': { ... },
        'preview_rows': [ ... ],  (ilk 5 satır)
        'total_rows': int,
        'detected_delimiter': str,
        'detected_encoding': str,
        'detected_program': str | None,
        'raw_headers': [str],
        'warnings': [str],
    }
    """
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext in ('xlsx', 'xls'):
        return _preview_excel(content, filename)
    else:
        return _preview_csv(content, filename)


def _preview_csv(content: bytes, filename: str) -> Dict:
    """CSV dosyası için preview oluştur."""
    warnings: List[str] = []

    # Encoding tespiti
    text = None
    detected_encoding = 'unknown'
    for encoding in ['utf-8-sig', 'utf-8', 'cp1254', 'latin-1']:
        try:
            text = content.decode(encoding)
            detected_encoding = encoding
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        raise ValueError("Dosya karakter kodlaması tanınamadı")

    # Ayırıcı tespiti (daha akıllı)
    lines = text.strip().split('\n')
    if not lines:
        raise ValueError("Dosya boş")

    first_line = lines[0]
    delimiter = _detect_delimiter(first_line)

    # Header parse
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_raw: List[List[str]] = []
    for row in reader:
        rows_raw.append(row)

    if not rows_raw:
        raise ValueError("Dosyada satır bulunamadı")

    raw_headers = [h.strip() for h in rows_raw[0]]

    # Sütun eşleştirme
    column_mapping = smart_map_columns(raw_headers)

    # Program tespiti
    detected_program = detect_accounting_program(raw_headers, delimiter)

    # Zorunlu alan kontrolü
    if 'hesap_kodu' not in column_mapping:
        warnings.append("'hesap_kodu' sütunu tespit edilemedi. Manuel eşleştirme gerekli.")
    if 'bakiye' not in column_mapping:
        warnings.append("'bakiye' sütunu tespit edilemedi. Manuel eşleştirme gerekli.")

    # Preview satırları (ilk 5 veri satırı)
    preview_rows = []
    for row in rows_raw[1:6]:  # header hariç ilk 5
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(raw_headers):
                row_dict[raw_headers[i]] = val.strip()
        preview_rows.append(row_dict)

    total_rows = len(rows_raw) - 1  # header hariç

    return {
        'column_mapping': column_mapping,
        'preview_rows': preview_rows,
        'total_rows': total_rows,
        'detected_delimiter': delimiter,
        'detected_encoding': detected_encoding,
        'detected_program': detected_program,
        'raw_headers': raw_headers,
        'warnings': warnings,
    }


def _preview_excel(content: bytes, filename: str) -> Dict:
    """Excel dosyası için preview oluştur."""
    warnings: List[str] = []

    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel desteği için openpyxl gerekli. pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel dosyasında aktif sayfa bulunamadı")

    # Başlıkları oku
    raw_headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        raw_headers.append(str(cell.value or '').strip())

    # Sütun eşleştirme
    column_mapping = smart_map_columns(raw_headers)
    detected_program = detect_accounting_program(raw_headers, ';')

    if 'hesap_kodu' not in column_mapping:
        warnings.append("'hesap_kodu' sütunu tespit edilemedi. Manuel eşleştirme gerekli.")
    if 'bakiye' not in column_mapping:
        warnings.append("'bakiye' sütunu tespit edilemedi. Manuel eşleştirme gerekli.")

    # Preview satırları
    preview_rows = []
    total_rows = 0
    for row_cells in ws.iter_rows(min_row=2):
        total_rows += 1
        if total_rows <= 5:
            row_dict = {}
            cells = list(row_cells)
            for i, cell in enumerate(cells):
                if i < len(raw_headers):
                    row_dict[raw_headers[i]] = str(cell.value or '').strip()
            preview_rows.append(row_dict)

    wb.close()

    return {
        'column_mapping': column_mapping,
        'preview_rows': preview_rows,
        'total_rows': total_rows,
        'detected_delimiter': 'N/A (Excel)',
        'detected_encoding': 'N/A (Excel)',
        'detected_program': detected_program,
        'raw_headers': raw_headers,
        'warnings': warnings,
    }


def _detect_delimiter(first_line: str) -> str:
    """CSV ayırıcı tespiti: en çok kullanılan ayırıcıyı bul."""
    counts = {
        ';': first_line.count(';'),
        '\t': first_line.count('\t'),
        ',': first_line.count(','),
        '|': first_line.count('|'),
    }
    best = max(counts, key=counts.get)
    if counts[best] == 0:
        return ','  # fallback
    return best


def parse_ekstre_with_mapping(
    content: bytes,
    filename: str,
    column_mapping: Dict[str, Dict],
) -> List[Dict]:
    """
    Kullanıcı onaylı sütun haritasıyla ekstre parse eder.
    Preview'dan gelen (veya SMMM'nin düzelttiği) mapping kullanılır.

    Returns: List[{hesap_kodu, karsi_taraf, bakiye}]
    """
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext in ('xlsx', 'xls'):
        return _parse_excel_with_mapping(content, filename, column_mapping)
    else:
        return _parse_csv_with_mapping(content, filename, column_mapping)


def _parse_csv_with_mapping(
    content: bytes,
    filename: str,
    column_mapping: Dict[str, Dict],
) -> List[Dict]:
    """CSV'yi onaylı mapping ile parse et."""
    rows = []

    # Encoding
    text = None
    for encoding in ['utf-8-sig', 'utf-8', 'cp1254', 'latin-1']:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Dosya karakter kodlaması tanınamadı")

    lines = text.strip().split('\n')
    if not lines:
        raise ValueError("Dosya boş")

    delimiter = _detect_delimiter(lines[0])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_raw = list(reader)

    if len(rows_raw) < 2:
        raise ValueError("Dosyada veri satırı yok")

    raw_headers = [h.strip() for h in rows_raw[0]]

    # Mapping'den index bul
    hesap_idx = _get_mapped_index(column_mapping, 'hesap_kodu', raw_headers)
    karsi_idx = _get_mapped_index(column_mapping, 'karsi_taraf', raw_headers)
    bakiye_idx = _get_mapped_index(column_mapping, 'bakiye', raw_headers)
    borc_idx = _get_mapped_index(column_mapping, 'borc', raw_headers)
    alacak_idx = _get_mapped_index(column_mapping, 'alacak', raw_headers)

    is_computed_bakiye = (
        bakiye_idx is None
        and column_mapping.get('bakiye', {}).get('method') == 'computed'
    )

    if hesap_idx is None:
        raise ValueError("hesap_kodu sütunu eşleştirilemedi")

    for row in rows_raw[1:]:
        if hesap_idx >= len(row):
            continue
        hesap_kodu = row[hesap_idx].strip()
        if not hesap_kodu:
            continue

        karsi_taraf = ''
        if karsi_idx is not None and karsi_idx < len(row):
            karsi_taraf = row[karsi_idx].strip()

        if is_computed_bakiye and borc_idx is not None and alacak_idx is not None:
            b = _parse_turkish_number(row[borc_idx].strip() if borc_idx < len(row) else '0')
            a = _parse_turkish_number(row[alacak_idx].strip() if alacak_idx < len(row) else '0')
            bakiye = b - a
        elif bakiye_idx is not None and bakiye_idx < len(row):
            bakiye = _parse_turkish_number(row[bakiye_idx].strip())
        else:
            bakiye = 0.0

        rows.append({
            'hesap_kodu': hesap_kodu,
            'karsi_taraf': karsi_taraf,
            'bakiye': bakiye,
        })

    logger.info(f"[CariMutabakat] Smart CSV parse: {len(rows)} satır ({filename})")
    return rows


def _parse_excel_with_mapping(
    content: bytes,
    filename: str,
    column_mapping: Dict[str, Dict],
) -> List[Dict]:
    """Excel'i onaylı mapping ile parse et."""
    rows = []
    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel desteği için openpyxl gerekli")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel'de aktif sayfa yok")

    raw_headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        raw_headers.append(str(cell.value or '').strip())

    hesap_idx = _get_mapped_index(column_mapping, 'hesap_kodu', raw_headers)
    karsi_idx = _get_mapped_index(column_mapping, 'karsi_taraf', raw_headers)
    bakiye_idx = _get_mapped_index(column_mapping, 'bakiye', raw_headers)
    borc_idx = _get_mapped_index(column_mapping, 'borc', raw_headers)
    alacak_idx = _get_mapped_index(column_mapping, 'alacak', raw_headers)

    is_computed_bakiye = (
        bakiye_idx is None
        and column_mapping.get('bakiye', {}).get('method') == 'computed'
    )

    if hesap_idx is None:
        wb.close()
        raise ValueError("hesap_kodu sütunu eşleştirilemedi")

    for row_cells in ws.iter_rows(min_row=2):
        cells = list(row_cells)
        if hesap_idx >= len(cells):
            continue
        hesap_kodu = str(cells[hesap_idx].value or '').strip()
        if not hesap_kodu:
            continue

        karsi_taraf = ''
        if karsi_idx is not None and karsi_idx < len(cells):
            karsi_taraf = str(cells[karsi_idx].value or '').strip()

        if is_computed_bakiye and borc_idx is not None and alacak_idx is not None:
            b_val = cells[borc_idx].value if borc_idx < len(cells) else 0
            a_val = cells[alacak_idx].value if alacak_idx < len(cells) else 0
            b = float(b_val) if isinstance(b_val, (int, float)) else _parse_turkish_number(str(b_val or '0'))
            a = float(a_val) if isinstance(a_val, (int, float)) else _parse_turkish_number(str(a_val or '0'))
            bakiye = b - a
        elif bakiye_idx is not None and bakiye_idx < len(cells):
            bval = cells[bakiye_idx].value
            if isinstance(bval, (int, float)):
                bakiye = float(bval)
            else:
                bakiye = _parse_turkish_number(str(bval or '0'))
        else:
            bakiye = 0.0

        rows.append({
            'hesap_kodu': hesap_kodu,
            'karsi_taraf': karsi_taraf,
            'bakiye': bakiye,
        })

    wb.close()
    logger.info(f"[CariMutabakat] Smart Excel parse: {len(rows)} satır ({filename})")
    return rows


def _get_mapped_index(
    column_mapping: Dict[str, Dict],
    field: str,
    raw_headers: List[str],
) -> Optional[int]:
    """Mapping'deki source_column'u raw_headers içinde bul."""
    info = column_mapping.get(field)
    if not info:
        return None

    # source_index doğrudan varsa ve geçerliyse kullan
    idx = info.get('source_index')
    if isinstance(idx, int) and idx >= 0 and idx < len(raw_headers):
        return idx

    # source_column ile aramaya devam
    source_col = info.get('source_column', '')
    for i, h in enumerate(raw_headers):
        if h.strip() == source_col.strip():
            return i
    return None
